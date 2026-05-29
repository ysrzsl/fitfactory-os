"""报表导出服务（Excel）"""
import io
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session


def _style_header(ws, row: int, cols: int):
    """设置表头样式"""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def export_orders(db: Session) -> io.BytesIO:
    """导出订单列表为 Excel"""
    from src.models.order import Order
    wb = Workbook()
    ws = wb.active; ws.title = "订单列表"

    headers = ["订单号", "客户", "款号", "件数", "交期", "产线", "开工", "完工", "状态", "优先级"]
    for c, h in enumerate(headers, 1): ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    orders = db.query(Order).order_by(Order.delivery_date.asc()).all()
    for r, o in enumerate(orders, 2):
        ws.cell(row=r, column=1, value=o.order_number)
        ws.cell(row=r, column=2, value=o.customer_name)
        ws.cell(row=r, column=3, value=o.style_code)
        ws.cell(row=r, column=4, value=o.total_quantity)
        ws.cell(row=r, column=5, value=str(o.delivery_date))
        ws.cell(row=r, column=6, value=o.assigned_line)
        ws.cell(row=r, column=7, value=str(o.start_date) if o.start_date else "")
        ws.cell(row=r, column=8, value=str(o.end_date) if o.end_date else "")
        ws.cell(row=r, column=9, value=o.status)
        ws.cell(row=r, column=10, value=o.priority)

    _auto_width(ws)
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output


def export_payroll(db: Session, year: int, month: int) -> io.BytesIO:
    """导出工资表"""
    from src.services.payroll import calc_monthly_payroll
    workers = calc_monthly_payroll(year, month, db)

    wb = Workbook()
    ws = wb.active; ws.title = f"{year}年{month}月工资"

    headers = ["排名", "姓名", "工号", "总件数", "总工资", "偏离均值%", "异常"]
    for c, h in enumerate(headers, 1): ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for r, w in enumerate(workers, 2):
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=w["worker_name"])
        ws.cell(row=r, column=3, value=w.get("worker_id", ""))
        ws.cell(row=r, column=4, value=w["total_quantity"])
        ws.cell(row=r, column=5, value=round(w["total_pay"], 2))
        ws.cell(row=r, column=6, value=f"{w.get('deviation_from_avg_pct', 0)}%")
        ws.cell(row=r, column=7, value="是" if w.get("anomaly") else "否")

    _auto_width(ws)
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output


def export_materials(db: Session) -> io.BytesIO:
    """导出物料库存"""
    from src.models.material import Material
    wb = Workbook()
    ws = wb.active; ws.title = "物料库存"

    headers = ["编码", "名称", "类别", "库存", "单位", "安全库存", "状态", "供应商"]
    for c, h in enumerate(headers, 1): ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    materials = db.query(Material).all()
    for r, m in enumerate(materials, 2):
        ws.cell(row=r, column=1, value=m.material_code)
        ws.cell(row=r, column=2, value=m.material_name)
        ws.cell(row=r, column=3, value=m.category)
        ws.cell(row=r, column=4, value=m.current_stock)
        ws.cell(row=r, column=5, value=m.unit)
        ws.cell(row=r, column=6, value=m.safety_stock)
        ws.cell(row=r, column=7, value="缺料" if m.current_stock < m.safety_stock else "正常")
        ws.cell(row=r, column=8, value=m.supplier_name)

    _auto_width(ws)
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output
